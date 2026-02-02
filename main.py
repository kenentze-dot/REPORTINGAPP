import os
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.spinner import Spinner
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.popup import Popup
from kivy.uix.camera import Camera
from kivy.utils import platform
from openpyxl import load_workbook

class EnercapitaDDRApp(App):
    def build(self):
        self.selected_file = None
        self.layout = BoxLayout(orientation='vertical', padding=20, spacing=10)
        
        # Header
        self.layout.add_widget(Label(text="ENERCAPITA PASON SCANNER", font_size='20sp', size_hint_y=0.1))

        # File Status
        self.file_label = Label(text="No File Selected", color=(1,0,0,1), size_hint_y=0.05)
        self.layout.add_widget(self.file_label)

        # Browse Button
        btn_browse = Button(text="1. SELECT EXCEL FILE", size_hint_y=0.1)
        btn_browse.bind(on_press=self.open_file_browser)
        self.layout.add_widget(btn_browse)
        
        # Day Selection
        self.day_spinner = Spinner(
            text='2. SELECT REPORT DAY', 
            values=[f'Day ({i})' for i in range(1, 51)], 
            size_hint_y=0.1
        )
        self.layout.add_widget(self.day_spinner)
        
        # Camera
        self.camera = Camera(resolution=(1280, 720), play=True)
        self.layout.add_widget(self.camera)
        
        # Run Button
        btn_run = Button(text="3. SCAN & SAVE TO ROW 50", size_hint_y=0.15, background_color=(0, 0.7, 0.4, 1))
        btn_run.bind(on_press=self.capture_and_process)
        self.layout.add_widget(btn_run)
        
        return self.layout

    def open_file_browser(self, instance):
        path = '/sdcard/Documents/Reports' if platform == 'android' else '.'
        content = FileChooserIconView(path=path, filters=['*.xlsm'])
        self.popup = Popup(title="Select Rig File", content=content, size_hint=(0.9, 0.9))
        content.bind(on_submit=self.set_file)
        self.popup.open()

    def set_file(self, instance, selection, touch):
        if selection:
            self.selected_file = selection[0]
            self.file_label.text = f"FILE: {os.path.basename(self.selected_file)}"
            self.file_label.color = (0, 1, 0, 1)
            self.popup.dismiss()

    def capture_and_process(self, instance):
        if not self.selected_file or "SELECT" in self.day_spinner.text:
            return
        
        # Capture photo
        self.camera.export_to_png("scan.png")
        
        # Write to Excel
        wb = load_workbook(self.selected_file, keep_vba=True)
        ws = wb[self.day_spinner.text]
        
        # Sample Data Injection starting at Row 50
        ws['A50'] = "08:00"
        ws['B50'] = "12:00"
        ws['D50'] = "DRIVE PILE / DRILLING"
        
        wb.save(self.selected_file)
        self.file_label.text = "SUCCESS: ROW 50 UPDATED!"

if __name__ == '__main__':
    EnercapitaDDRApp().run()
